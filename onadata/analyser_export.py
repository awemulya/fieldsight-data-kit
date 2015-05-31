'''
Created on May 14, 2015

@author: esmail
'''


import tempfile
import io
import shutil
import copy
from zipfile import (
    ZipFile,
    ZIP_DEFLATED,
)

import lxml.etree as etree
import xlrd
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook


NAMESPACES= {'xmlns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}


# Adapted from http://stackoverflow.com/a/9919409/1877326
def xls_as_xlsx(xls_file):
    # first open using xlrd
    source_workbook = xlrd.open_workbook(file_contents=xls_file.read())

    # Create the destination workbook, deleting and auto-generated worksheets.
    destination_workbook = openpyxl.Workbook() # TODO: Would like to figure out how to make appends work with a "write_only" workbook.
    for wksht_nm in destination_workbook.sheetnames:
        worksheet= destination_workbook.get_sheet_by_name(wksht_nm)
        destination_workbook.remove_sheet(worksheet)

    worksheet_names= ['survey', 'choices']
    for wksht_nm in worksheet_names:
        source_worksheet= source_workbook.sheet_by_name(wksht_nm)
        destination_worksheet= destination_workbook.create_sheet(title=wksht_nm)

        for row in xrange(source_worksheet.nrows):
            destination_worksheet.append( (source_worksheet.cell_value(row, col) for col in xrange(source_worksheet.ncols)) )

    return io.BytesIO(save_virtual_workbook(destination_workbook))

def copy_cells(source_worksheet_file, destination_worksheet_file_path, new_string_indices):
    destination_worksheet= etree.parse(destination_worksheet_file_path)

    namespace_prefix= '{' + NAMESPACES['xmlns'] + '}'
    for _, source_dimension in etree.iterparse(source_worksheet_file, tag=namespace_prefix+'dimension'):
        destination_dimension= destination_worksheet.xpath('.//xmlns:dimension', namespaces=NAMESPACES)[0]
        destination_dimension.attrib['ref']= source_dimension.attrib['ref']
        source_dimension.getroottree().getroot().clear() # FIXME: Necessary? Enough?
        break
    else:
        raise ValueError('No "dimension" element found in source data file.')
    source_worksheet_file.seek(0)

    # Copy the data over row by row, iterating through the (potentially large)
    #   source instead of loading at once.
    destination_sheetData_element= destination_worksheet.xpath('//xmlns:sheetData', namespaces=NAMESPACES)[0]
    for _, source_row in etree.iterparse(source_worksheet_file, tag=namespace_prefix+'row'):
        # Create a new row element.
        destination_row= etree.Element(namespace_prefix+'row', NAMESPACES)
        destination_row.attrib.update(source_row.attrib)

        # Copy over the cells one by one.
        for source_c in source_row:
            destination_c= copy.deepcopy(source_c)
            # Remap references to shared strings.
            if destination_c.attrib['t'] == 's':
                values= destination_c.xpath('.//xmlns:v', namespaces=NAMESPACES)
                if values:
                    destination_v= values[0]
                    destination_v.text= unicode(new_string_indices[int(destination_v.text)])
            destination_row.append(destination_c)

        # Clean up elements of the source worksheet to save memory.
        while source_row.getprevious():
            source_row.getprevious().clean()
            print 'Previous row cleaned'
        source_row.clear()

        # Append in the copied row.
        destination_sheetData_element.append(destination_row)

    # Save the changes.
    with open(destination_worksheet_file_path, 'w') as destination_worksheet_file:
        destination_worksheet.write(destination_worksheet_file, encoding='UTF-8')

def splice_shared_strings(source_file, destination_etree, new_string_indices):
    source_etree= etree.parse(source_file)
    original_string_map= dict()
    destination_root= destination_etree.getroot()
    for i, t_element in enumerate(destination_root.iterfind('.//xmlns:t', NAMESPACES)):
        original_string_map[t_element.text]= i

    uniqueCount= len(original_string_map)
    for i, t_element in enumerate(source_etree.iterfind('//xmlns:t', NAMESPACES)):
        text= t_element.text
        if text in original_string_map:
            new_string_indices[i]= original_string_map[text]
        else:
            # Copy the "si" element over to `destination_etree`.
            si_element= t_element.getparent()
            destination_root.append(si_element)
            new_string_indices[i]= uniqueCount
            uniqueCount+= 1

    # Update the "uniqueCount" attribute.
    destination_root.attrib['uniqueCount']= unicode(uniqueCount)

def insert_xlsform_worksheets(analyser_shared_strings, analyser_survey_worksheet_file_path, analyser_choices_worksheet_file_path, survey_file_xls):
    # Create an XLSX copy of the survey file.
    survey_file_xlsx= xls_as_xlsx(survey_file_xls)

    with ZipFile(survey_file_xlsx) as survey_zipfile:
        # Splice over the shared strings.
        with survey_zipfile.open('xl/sharedStrings.xml') as shared_strings_file:
            new_string_indices= dict()
            splice_shared_strings(shared_strings_file, analyser_shared_strings, new_string_indices)
        # Copy over the "survey" sheet.
        with survey_zipfile.open('xl/worksheets/sheet1.xml') as source_survey_worksheet_file:
            # Create a tempfile that supports seeking.
            with tempfile.TemporaryFile('w+') as source_survey_worksheet_tempfile:
                source_survey_worksheet_tempfile.write(source_survey_worksheet_file.read())
                source_survey_worksheet_tempfile.seek(0)
                copy_cells(source_survey_worksheet_tempfile, analyser_survey_worksheet_file_path, new_string_indices)
        # Copy over the "choices" sheet.
        with survey_zipfile.open('xl/worksheets/sheet2.xml') as source_choices_worksheet_file:
            # Create a tempfile that supports seeking.
            with tempfile.TemporaryFile('w+') as source_choices_worksheet_tempfile:
                source_choices_worksheet_tempfile.write(source_choices_worksheet_file.read())
                source_choices_worksheet_tempfile.seek(0)
                copy_cells(source_choices_worksheet_tempfile, analyser_choices_worksheet_file_path, new_string_indices)

def insert_data_sheet(analyser_shared_strings, analyser_data_sheet_file_path, data_file_xlsx):
    with ZipFile(data_file_xlsx) as data_zipfile:
        with data_zipfile.open('xl/sharedStrings.xml') as shared_strings_file:
            new_string_indices= dict()
            splice_shared_strings(shared_strings_file, analyser_shared_strings, new_string_indices)
        # FIXME: Not safe for multi-sheet data such for repeating groups.
        with data_zipfile.open('xl/worksheets/sheet1.xml') as data_worksheet_file:
            with tempfile.TemporaryFile('w+') as data_worksheet_tempfile:
                data_worksheet_tempfile.write(data_worksheet_file.read())
                data_worksheet_tempfile.seek(0)
                copy_cells(data_worksheet_tempfile, analyser_data_sheet_file_path, new_string_indices)

def generate_analyser(analyser_file_xlsx, data_file_xlsx, survey_file_xls):
    # Unzip the analyser in preparation for modifications.
    temp_directory_path= tempfile.mkdtemp(prefix='analyser_temp_')
    try:
        with ZipFile(analyser_file_xlsx) as analyser_zipfile:
            analyser_zipfile.extractall(temp_directory_path)
            zip_contents= [f.filename for f in analyser_zipfile.filelist]

        analyser_shared_strings= etree.parse(temp_directory_path + '/xl/sharedStrings.xml')

        # Copy over the XLSForm's "survey" and "choices" (if present) worksheets.
        survey_worksheet_file_path= temp_directory_path + '/xl/worksheets/sheet9.xml'
        choices_worksheet_file_path= temp_directory_path + '/xl/worksheets/sheet10.xml'
        insert_xlsform_worksheets(analyser_shared_strings, survey_worksheet_file_path, choices_worksheet_file_path, survey_file_xls)

        # Copy the data over to the analyser.
        analyser_data_sheet_file_path= temp_directory_path + '/xl/worksheets/sheet8.xml'
        insert_data_sheet(analyser_shared_strings, analyser_data_sheet_file_path, data_file_xlsx)

        # Finalize the changes for export.
        with open(temp_directory_path + '/xl/sharedStrings.xml', 'wb') as analyser_shared_strings_file:
            analyser_shared_strings.write(analyser_shared_strings_file, encoding='UTF-8')
        xlsx_out= io.BytesIO()
        with ZipFile(xlsx_out, 'w', compression=ZIP_DEFLATED) as zipfile_out:
            for file_path in zip_contents:
                zipfile_out.write(temp_directory_path + '/' + file_path, file_path)

    finally:
        # Clean up temporary files.
        shutil.rmtree(temp_directory_path)

    xlsx_out.seek(0)
    return xlsx_out

if __name__ == '__main__':
    analyser_file_xlsx= '/home/esmail/Downloads/KoBoToolbox_Excel_Data_Analyser_1.23_TEMPLATE.xlsx'
    data_file_xlsx= '/home/esmail/Downloads/uga_14_v6_2015_03_02_14_34_38.xlsx'
    survey_file= open('/home/esmail/Downloads/uga_14_v6.xls')
    xlsx_out= generate_analyser(analyser_file_xlsx, data_file_xlsx, survey_file)
    with open('/home/esmail/Downloads/generated_analyser.xlsx', 'wb') as f:
        f.write(xlsx_out.read())
